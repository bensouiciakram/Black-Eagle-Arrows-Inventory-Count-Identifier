"""
Black Eagle Arrows Inventory Count Identifier

This project automates the process of scraping information from blackeaglearrows.com
and gathering the quantity of inventory for each product. The automation script uses
Playwright for web scraping and Jinja2 for generating reports about the results.

The scraper extracts product information, identifies inventory quantities by attempting
to add products to cart, and generates comprehensive reports with inventory details.

Author: Project Developer
License: MIT
"""

import pandas as pd 
import logging 
from pathlib import Path 
from playwright.async_api import async_playwright, TimeoutError
from playwright.async_api._generated import Browser
from playwright.async_api._generated import BrowserContext
from playwright.async_api._generated import Page
from playwright.async_api._generated import ElementHandle
from playwright._impl._api_types import Error
from jinja2 import Environment, FileSystemLoader
import asyncio
from parsel import Selector 
from typing import List, Dict, Any, Set, Optional, Tuple
from itertools import product 
from copy import deepcopy 
import sys 
from datetime import datetime 
from re import sub, compile
import pickle 
import os 
import openpyxl 
from openpyxl.worksheet.hyperlink import Hyperlink
from json.decoder import JSONDecodeError

# Global configuration
PAGES_NUMBER = int(input('How many tabs do you want: '))
HEADLESS_SIGN = input('Enter y if you want to see the browser and n if not: ')
HEADLESS = False if HEADLESS_SIGN == 'y' else True

# Columns that uniquely identify a production variation
COLS = [
    'SKU',
    '1DroplistDesc',
    '1DroplistValue',
    '2DroplistDesc',
    '2DroplistValue',
    '3DroplistDesc',
    '3DroplistValue',
    '4DroplistDesc',
    '4DroplistValue'
]

# List of all available listing pages on the site
LISTING_URLS = [
    'https://blackeaglearrows.com/bow-accessories/',
    'https://blackeaglearrows.com/arrows/',
    'https://blackeaglearrows.com/components/',
    'https://blackeaglearrows.com/gear/'
]

# Characters to clean from file names
NAME_CLEANING_LIST = ['/', '"']

# Global data containers
data: List[Dict[str, Any]] = []
failed_urls: Set[str] = set()
products_urls: Set[str] = set()
logger: logging.Logger = None

def create_logger(logging_level: int) -> logging.RootLogger:
    """
    Create and configure a logger with both console and file handlers.
    
    Args:
        logging_level: Logging level (e.g., logging.DEBUG, logging.INFO)
        
    Returns:
        logging.RootLogger: Configured logger instance
    """
    Path(__file__).parent.joinpath('logs').mkdir(exist_ok=True)
    logger = logging.getLogger(__name__)
    logger.setLevel(logging_level)
    
    # Console handler
    stream_handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter('%(levelname)s:%(message)s')
    stream_handler.setFormatter(formatter)
    
    # File handler
    timestamp = str(datetime.now()).replace(":", "")
    file_handler = logging.FileHandler(f'logs/logs_{timestamp}.log')
    file_handler.setFormatter(formatter)
    
    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)
    return logger

def load_existing_data() -> None:
    """
    Load existing data from pickle file if available, otherwise initialize empty data.
    """
    global data
    if os.path.exists('data.pkl'):
        df = pd.DataFrame(pickle.load(open('data.pkl', 'rb')))
        data = df[~df[COLS].duplicated()].to_dict(orient='records')
        logger.info(f"Loaded {len(data)} existing records from data.pkl")
    else:
        data = []
        logger.info("No existing data found, starting fresh")

async def check_handled_url(page: Page) -> bool:
    """
    Check if the URL has been processed before by comparing product combinations.
    
    Args:
        page: Playwright page object
        
    Returns:
        bool: True if URL has been handled, False otherwise
    """
    global data
    attrs_dict = await get_options_dict(page)
    all_combinations = get_all_combinations(*(attrs_dict.values()))
    return len(list(all_combinations)) == len([item['product_name'] == page.url for item in data])

async def get_total_pages(page: Page) -> int:
    """
    Get the total number of pages for a listing page.
    
    Args:
        page: Playwright page object
        
    Returns:
        int: Total number of pages, defaults to 1 if pagination not found
    """
    selector = Selector(text=await page.content())
    try:
        pagination_numbers = selector.xpath('//a[contains(@class,"listing-pagination-link")]').re(r'\d+')
        return max([int(number) for number in pagination_numbers]) if pagination_numbers else 1
    except ValueError:
        logger.warning("Could not determine total pages, defaulting to 1")
        return 1

async def get_page_products_urls(page: Page) -> List[str]:
    """
    Extract all product URLs from a listing page.
    
    Args:
        page: Playwright page object
        
    Returns:
        List[str]: List of product URLs found on the page
    """
    selector = Selector(text=await page.content())
    return selector.xpath('//div[contains(@class,"product-item-image")]//a/@href').getall()

def format_date() -> str:
    """
    Format the current date as mm-dd-yyyy.
    
    Returns:
        str: Formatted date string
    """
    date = datetime.today()
    return f'{date.month}-{date.day}-{date.year}'

def clean_file_name(name: str) -> str:
    """
    Clean unwanted symbols from file name.
    
    Args:
        name: Original file name
        
    Returns:
        str: Cleaned file name with .html extension
    """
    for element in NAME_CLEANING_LIST:
        name = name.replace(element, ' ')
    return sub(r'\s+', ' ', name) + '.html'

def save_description(selector: Selector, product_item: Dict[str, Any]) -> str:
    """
    Save product description to HTML file.
    
    Args:
        selector: Parsel selector object
        product_item: Product data dictionary
        
    Returns:
        str: Absolute path to the saved description file
    """
    description_folder = Path(__file__).parent.joinpath('descriptions')
    description_folder.mkdir(exist_ok=True)
    description_path = description_folder.joinpath(clean_file_name(product_item['product_name']))
    
    description_content = selector.xpath('//section[@id="description"]').get()
    with open(description_path, 'w', encoding='utf-8') as file:
        file.write(description_content if description_content else '')
    
    return str(description_path.absolute())

async def get_product_item(page: Page) -> Dict[str, Any]:
    """
    Extract initial product information from a product page.
    
    Args:
        page: Playwright page object
        
    Returns:
        Dict[str, Any]: Dictionary containing product information
    """
    global data
    selector = Selector(text=await page.content())
    
    product_item = {
        'SKU': '',  # selector.xpath('string(//span[contains(text(),"SKU")]/following-sibling::span[1])').get().strip()
        'Brand': selector.xpath('string(//a[@class="product-brand"])').get().strip(),
        'product_name': selector.xpath('//h1/text()').get(),
        'URL': page.url,
        '1DroplistDesc': '',
        '1DroplistValue': '',
        '2DroplistDesc': '',
        '2DroplistValue': '',
        '3DroplistDesc': '',
        '3DroplistValue': '',
        '4DroplistDesc': '',
        '4DroplistValue': '',
        'Price': '',
        'Current stock': 0,
        'Current stock date': format_date(),
        'Previous stock': 0,
        'Previous stock date': '',
        'Description_path': '',
        'Description': '',
        'Item photo URL': selector.xpath('//img[contains(@class,"product-main-image-slide")]/@src').get()
    }
    
    # Save description and update paths
    product_item['Description_path'] = save_description(selector, product_item)
    product_item['Description'] = selector.xpath('//section[@id="description"]').get()
    product_item['Description'] = remove_hyperlinks(product_item)
    
    return product_item

def export() -> None:
    """
    Export the extracted data to CSV file.
    """
    global data
    output_folder = Path(__file__).parent.joinpath('outputs')
    output_folder.mkdir(exist_ok=True)
    
    df = pd.DataFrame(data)
    output_path = output_folder.joinpath('output.csv')
    df.to_csv(output_path, index=False)
    logger.info(f"Exported {len(data)} records to {output_path}")

async def handle_url(browser: Browser, url: str) -> None:
    """
    Handle the extraction logic for a single product URL.
    
    Args:
        browser: Playwright browser instance
        url: Product URL to process
    """
    global failed_urls, data
    
    context = await browser.new_context()
    page = await context.new_page()
    
    try:
        await page.goto(url, timeout=60000)
        
        if await check_handled_url(page):
            logger.info(f"URL already handled: {url}")
            return
            
        primary_item = await get_product_item(page)
        
        # Check if product is unavailable
        unavailable_button = await page.query_selector('//div[@class="product-add-to-cart form-field"]/input[@value="Unavailable"]')
        if unavailable_button:
            logger.info('No adding cart button available')
            logger.info(f'Item scraped: {primary_item["product_name"]}')
            data.append(primary_item)
        else:
            try:
                await inventory_identifier(page, primary_item)
            except TimeoutError:
                logger.error(f'Timeout problem in link: {url}')
                failed_urls.add(url)
                pickle.dump(failed_urls, open('failed_urls.pkl', 'wb'))
                
    except Exception as e:
        logger.error(f"Error processing URL {url}: {e}")
        failed_urls.add(url)
    finally:
        export()
        pickle.dump(data, open('data.pkl', 'wb'))
        await page.close()
        await context.close()

async def handle_listing(browser: Browser, listing_url: str) -> None:
    """
    Handle the listing page extraction logic to collect product URLs.
    
    Args:
        browser: Playwright browser instance
        listing_url: Listing page URL to process
    """
    global data, products_urls
    
    context = await browser.new_context()
    page = await context.new_page()
    
    try:
        await page.goto(listing_url, timeout=60000)
        total_pages = await get_total_pages(page)
        
        logger.info(f"Processing listing: {listing_url} with {total_pages} pages")
        
        for page_num in range(1, total_pages + 1):
            if page_num > 1:
                await page.goto(f"{listing_url}?page={page_num}", timeout=60000)
            
            product_urls = await get_page_products_urls(page)
            products_urls.update(product_urls)
            logger.info(f"Page {page_num}: Found {len(product_urls)} products")
            
    except Exception as e:
        logger.error(f"Error processing listing {listing_url}: {e}")
    finally:
        await page.close()
        await context.close()

async def get_attr_values(attr_handle: ElementHandle) -> List[str]:
    """
    Extract attribute values from an attribute handle.
    
    Args:
        attr_handle: Playwright element handle for attribute
        
    Returns:
        List[str]: List of attribute values
    """
    return await attr_handle.query_selector_all('option')

async def get_options_dict(page: Page) -> Dict[str, List[str]]:
    """
    Get dictionary of attribute names and their available values.
    
    Args:
        page: Playwright page object
        
    Returns:
        Dict[str, List[str]]: Dictionary mapping attribute names to their values
    """
    attr_handles = await page.query_selector_all('//select[contains(@class,"product-attribute-select")]')
    options_dict = {}
    
    for attr_handle in attr_handles:
        attr_name = await get_attr_name(attr_handle)
        attr_values = await get_attr_values(attr_handle)
        options_dict[attr_name] = [await value.get_attribute('value') for value in attr_values]
    
    return options_dict

async def select_attr_option(option_handle: ElementHandle, attr_value: str) -> None:
    """
    Select an attribute option by value.
    
    Args:
        option_handle: Playwright element handle for option
        attr_value: Value to select
    """
    await option_handle.select_option(attr_value)

def get_all_combinations(*lists: List[str]) -> List[Tuple[str, ...]]:
    """
    Generate all possible combinations of attribute values.
    
    Args:
        *lists: Variable number of attribute value lists
        
    Returns:
        List[Tuple[str, ...]]: All possible combinations
    """
    return list(product(*lists))

async def try_inventory_quantity(page: Page, check_value: int) -> bool:
    """
    Test if a specific quantity can be added to cart.
    
    Args:
        page: Playwright page object
        check_value: Quantity to test
        
    Returns:
        bool: True if quantity can be added, False otherwise
    """
    try:
        # Set quantity input
        quantity_input = await page.query_selector('//input[@name="quantity"]')
        if quantity_input:
            await quantity_input.fill(str(check_value))
        
        # Try to add to cart
        add_to_cart_button = await page.query_selector('//button[contains(@class,"add-to-cart")]')
        if add_to_cart_button:
            await add_to_cart_button.click()
            await page.wait_for_timeout(2000)
            
            # Check for error messages
            error_selectors = [
                '//div[contains(@class,"error")]',
                '//div[contains(@class,"alert")]',
                '//span[contains(text(),"not available")]'
            ]
            
            for selector in error_selectors:
                error_element = await page.query_selector(selector)
                if error_element:
                    return False
            
            return True
        
        return False
        
    except Exception as e:
        logger.error(f"Error testing inventory quantity {check_value}: {e}")
        return False

async def get_inventory_value(page: Page, check_value: int) -> int:
    """
    Get the actual inventory value by testing quantities.
    
    Args:
        page: Playwright page object
        check_value: Quantity to check
        
    Returns:
        int: Actual inventory quantity
    """
    if await try_inventory_quantity(page, check_value):
        return check_value
    else:
        # Binary search for actual inventory
        left, right = 1, check_value - 1
        while left <= right:
            mid = (left + right) // 2
            if await try_inventory_quantity(page, mid):
                left = mid + 1
            else:
                right = mid - 1
        return right

async def get_attr_name(attr_handle: ElementHandle) -> str:
    """
    Extract attribute name from attribute handle.
    
    Args:
        attr_handle: Playwright element handle for attribute
        
    Returns:
        str: Attribute name
    """
    return await attr_handle.get_attribute('name')

async def get_attr_value(page: Page, attr_value: str) -> str:
    """
    Get attribute value from page.
    
    Args:
        page: Playwright page object
        attr_value: Attribute value to extract
        
    Returns:
        str: Extracted attribute value
    """
    return attr_value

async def inventory_identifier(page: Page, primary_item: Dict[str, Any], guessed_initial_value: int = 100) -> int:
    """
    Identify inventory quantity for a product by testing cart additions.
    
    Args:
        page: Playwright page object
        primary_item: Primary product item dictionary
        guessed_initial_value: Initial guess for inventory quantity
        
    Returns:
        int: Identified inventory quantity
    """
    global data
    
    attrs_dict = await get_options_dict(page)
    all_combinations = get_all_combinations(*(attrs_dict.values()))
    
    for combination in all_combinations:
        # Select attributes
        attr_handles = await page.query_selector_all('//select[contains(@class,"product-attribute-select")]')
        for i, (attr_handle, value) in enumerate(zip(attr_handles, combination)):
            await select_attr_option(attr_handle, value)
            await page.wait_for_timeout(500)
        
        # Create product variation item
        variation_item = deepcopy(primary_item)
        for i, (attr_name, attr_value) in enumerate(zip(attrs_dict.keys(), combination)):
            variation_item[f'{i+1}DroplistDesc'] = attr_name
            variation_item[f'{i+1}DroplistValue'] = attr_value
        
        # Test inventory quantity
        inventory_quantity = await get_inventory_value(page, guessed_initial_value)
        variation_item['Current stock'] = inventory_quantity
        
        logger.info(f"Product variation: {variation_item['product_name']} - Stock: {inventory_quantity}")
        data.append(variation_item)
    
    return len(all_combinations)

def rename_columns(data_container: List[Dict[str, Any]], **cols: str) -> List[Dict[str, Any]]:
    """
    Rename columns in data container.
    
    Args:
        data_container: List of data dictionaries
        **cols: Column mapping (old_name=new_name)
        
    Returns:
        List[Dict[str, Any]]: Data with renamed columns
    """
    for item in data_container:
        for old_name, new_name in cols.items():
            if old_name in item:
                item[new_name] = item.pop(old_name)
    return data_container

def remove_hyperlinks(product_item: Dict[str, Any]) -> str:
    """
    Remove hyperlinks from product description.
    
    Args:
        product_item: Product item dictionary
        
    Returns:
        str: Description with hyperlinks removed
    """
    description = product_item.get('Description', '')
    if description:
        # Remove HTML links while keeping text
        description = sub(r'<a[^>]*>(.*?)</a>', r'\1', description)
    return description

def reshape_description(product_item: Dict[str, Any]) -> str:
    """
    Reshape product description for better formatting.
    
    Args:
        product_item: Product item dictionary
        
    Returns:
        str: Reshaped description
    """
    description = product_item.get('Description', '')
    if description:
        # Clean up HTML and format
        description = sub(r'<[^>]+>', '', description)  # Remove HTML tags
        description = sub(r'\s+', ' ', description)     # Normalize whitespace
        description = description.strip()
    return description

def standardize_data() -> Dict[str, Any]:
    """
    Standardize the collected data for reporting.
    
    Returns:
        Dict[str, Any]: Standardized data structure
    """
    global data
    
    if not data:
        return {
            'total_products': 0,
            'products_with_stock': 0,
            'products_out_of_stock': 0,
            'total_inventory': 0,
            'categories': {},
            'low_stock_products': [],
            'out_of_stock_products': []
        }
    
    df = pd.DataFrame(data)
    
    # Calculate statistics
    total_products = len(df)
    products_with_stock = len(df[df['Current stock'] > 0])
    products_out_of_stock = len(df[df['Current stock'] == 0])
    total_inventory = df['Current stock'].sum()
    
    # Category analysis
    categories = {}
    if 'Brand' in df.columns:
        categories = df.groupby('Brand')['Current stock'].agg(['sum', 'count']).to_dict('index')
    
    # Low stock products (less than 10 items)
    low_stock_products = df[df['Current stock'] < 10][['product_name', 'Current stock', 'Brand']].to_dict('records')
    
    # Out of stock products
    out_of_stock_products = df[df['Current stock'] == 0][['product_name', 'Brand']].to_dict('records')
    
    return {
        'total_products': total_products,
        'products_with_stock': products_with_stock,
        'products_out_of_stock': products_out_of_stock,
        'total_inventory': total_inventory,
        'categories': categories,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'data': data
    }

def generate_report(template_name: str, data: List[Dict[str, Any]]) -> None:
    """
    Generate HTML report using Jinja2 template.
    
    Args:
        template_name: Name of the template file
        data: Data to include in the report
    """
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template(template_name)
    
    standardized_data = standardize_data()
    report_content = template.render(**standardized_data)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_path = f'reports/report_{timestamp}.html'
    
    Path('reports').mkdir(exist_ok=True)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    logger.info(f"Report generated: {report_path}")

def create_hypertext() -> None:
    """
    Create hypertext links for the output Excel file.
    """
    output_folder = Path(__file__).parent.joinpath('outputs')
    excel_path = output_folder.joinpath('output_final.xlsx')
    
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Add hyperlinks to description files
        for row in range(2, ws.max_row + 1):
            description_path = ws.cell(row=row, column=ws.max_column).value
            if description_path and os.path.exists(description_path):
                ws.cell(row=row, column=ws.max_column).hyperlink = description_path
        
        wb.save(excel_path)
        logger.info("Hyperlinks added to Excel file")

def update_output() -> None:
    """
    Update the final output Excel file with processed data.
    """
    global data
    
    if not data:
        logger.warning("No data to export")
        return
    
    df = pd.DataFrame(data)
    
    # Clean and reshape descriptions
    df['Description'] = df['Description'].apply(lambda x: reshape_description({'Description': x}) if x else '')
    
    # Rename columns for better readability
    df = rename_columns(df.to_dict('records'), 
                       product_name='Product Name',
                       Current_stock='Current Stock',
                       Previous_stock='Previous Stock')
    
    # Export to Excel
    output_folder = Path(__file__).parent.joinpath('outputs')
    output_folder.mkdir(exist_ok=True)
    
    excel_path = output_folder.joinpath('output_final.xlsx')
    df_export = pd.DataFrame(df)
    df_export.to_excel(excel_path, index=False)
    
    logger.info(f"Final output exported to: {excel_path}")

async def gather_with_concurrency(n: int, *tasks) -> List[Any]:
    """
    Run tasks with limited concurrency.
    
    Args:
        n: Maximum number of concurrent tasks
        *tasks: Tasks to execute
        
    Returns:
        List[Any]: Results from all tasks
    """
    semaphore = asyncio.Semaphore(n)
    
    async def sem_task(task):
        async with semaphore:
            return await task
    
    return await asyncio.gather(*(sem_task(task) for task in tasks))

async def run(p) -> None:
    """
    Main execution function for the scraper.
    
    Args:
        p: Playwright instance
    """
    global HEADLESS, PAGES_NUMBER
    
    browser = await p.chromium.launch(headless=HEADLESS)
    
    # First, collect all product URLs from listing pages
    listing_tasks = [handle_listing(browser, url) for url in LISTING_URLS]
    await gather_with_concurrency(2, *listing_tasks)
    
    logger.info(f"Collected {len(products_urls)} product URLs")
    
    # Then process each product URL
    product_tasks = [handle_url(browser, url) for url in products_urls]
    await gather_with_concurrency(PAGES_NUMBER, *product_tasks)
    
    await browser.close()

async def main() -> None:
    """
    Main entry point for the Black Eagle Arrows scraper.
    """
    global logger
    
    # Initialize logger
    logger = create_logger(logging.DEBUG)
    logger.info("Starting Black Eagle Arrows Inventory Scraper")
    
    # Load existing data
    load_existing_data()
    
    try:
        async with async_playwright() as p:
            await run(p)
        
        # Generate final outputs
        update_output()
        create_hypertext()
        
        # Generate report
        if os.path.exists('template.txt'):
            generate_report('template.txt', data)
        
        logger.info("Scraping completed successfully")
        
    except Exception as e:
        logger.error(f"Scraping failed: {e}")
        raise

if __name__ == '__main__':
    asyncio.run(main())
